"""Per-glacier analysis: does local SMB deficit drive discharge acceleration?

Tests the hypothesis that cumulative SMB-driven thinning at individual
marine-terminating glaciers feeds back on discharge through changes in
gravitational driving stress.

If the feedback exists, glaciers with larger cumulative SMB deficits
should show stronger discharge acceleration.

Usage
-----
    from diagnostics.smb_discharge_feedback import analyze_smb_discharge_feedback
    fig = analyze_smb_discharge_feedback(
        'path/to/mouginot2019_data.xlsx',
        savefig='path/to/output.png',
    )
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import stats


def _read_mouginot_per_glacier(filepath):
    """Read per-glacier SMB and D from Mouginot 2019 Excel file.

    Returns
    -------
    dict with keys:
        'names'   : list of glacier names (TW only)
        'regions' : list of region codes
        'areas'   : (n_glaciers,) km²
        'smb'     : (n_glaciers, n_years_smb) Gt/yr per glacier
        'smb_years': (n_years_smb,)
        'd'       : (n_glaciers, n_years_d) Gt/yr per glacier
        'd_years' : (n_years_d,)
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── Sheet 5: Discharge per glacier ──
    ws_d = wb['(5) year_D_R23p2-5.5km']
    rows_d = list(ws_d.iter_rows(min_row=1, values_only=True))
    header_d = rows_d[2]  # row 3: NAME, REGION, LAT, LON, AREA, ...

    # Find year columns (numeric values > 1900 in header)
    d_year_cols = []
    d_years = []
    for j, val in enumerate(header_d):
        if isinstance(val, (int, float)) and 1900 < val < 2100:
            d_year_cols.append(j)
            d_years.append(int(val))
    d_years = np.array(d_years)

    # Extract TW glacier data
    tw_names = []
    tw_regions = []
    tw_areas = []
    tw_d = []
    tw_row_indices = []

    for i, row in enumerate(rows_d[3:], start=3):
        if row[8] == 'TW':  # column 8 = 'LT or TW'
            name = row[0]
            tw_names.append(name)
            tw_regions.append(row[1])
            area = row[4]
            tw_areas.append(float(area) if area is not None else np.nan)
            d_vals = []
            for j in d_year_cols:
                v = row[j]
                d_vals.append(float(v) if v is not None else np.nan)
            tw_d.append(d_vals)
            tw_row_indices.append(i)

    tw_d = np.array(tw_d)
    tw_areas = np.array(tw_areas)

    # ── Sheet 6: SMB per glacier ──
    ws_smb = wb['(6) SMB_R23p2-5km_Gl']
    rows_smb = list(ws_smb.iter_rows(min_row=1, values_only=True))

    # Row 4 (index 3) has center years
    smb_year_row = rows_smb[3]
    smb_year_cols = []
    smb_years = []
    for j, val in enumerate(smb_year_row):
        if isinstance(val, (int, float)) and 1900 < val < 2100:
            smb_year_cols.append(j)
            smb_years.append(int(val))
    smb_years = np.array(smb_years)

    # Build name→SMB mapping
    smb_by_name = {}
    for row in rows_smb[4:]:
        name = row[0]
        if name is None:
            continue
        vals = []
        for j in smb_year_cols:
            v = row[j]
            vals.append(float(v) if v is not None else np.nan)
        smb_by_name[name] = np.array(vals)

    # Match TW glaciers to SMB
    tw_smb = []
    matched_names = []
    matched_regions = []
    matched_areas = []
    matched_d = []

    for k, name in enumerate(tw_names):
        if name in smb_by_name:
            tw_smb.append(smb_by_name[name])
            matched_names.append(name)
            matched_regions.append(tw_regions[k])
            matched_areas.append(tw_areas[k])
            matched_d.append(tw_d[k])

    tw_smb = np.array(tw_smb)
    matched_d = np.array(matched_d)
    matched_areas = np.array(matched_areas)

    wb.close()

    return {
        'names': matched_names,
        'regions': matched_regions,
        'areas': matched_areas,
        'smb': tw_smb,
        'smb_years': smb_years,
        'd': matched_d,
        'd_years': d_years,
    }


def analyze_smb_discharge_feedback(
    mouginot_path,
    period_early=(1972, 1990),
    period_late=(2005, 2018),
    savefig=None,
    figsize=(16, 10),
):
    """Test whether per-glacier SMB deficit correlates with discharge change.

    For each marine-terminating glacier:
    - Cumulative SMB deficit = sum of (SMB - mean_early_SMB) over the record
    - Discharge change = mean(D_late) - mean(D_early)
    - Discharge acceleration from linear fit of D vs time

    Parameters
    ----------
    mouginot_path : str
        Path to mouginot2019_data.xlsx.
    period_early : tuple
        (start, end) years for early-period baseline.
    period_late : tuple
        (start, end) years for late-period comparison.
    savefig : str or None
        If provided, save figure to this path.
    figsize : tuple

    Returns
    -------
    fig : matplotlib Figure
    results : dict
        Summary statistics and per-glacier arrays.
    """
    data = _read_mouginot_per_glacier(mouginot_path)
    n_gl = len(data['names'])

    # Common years for SMB and D
    smb_years = data['smb_years']
    d_years = data['d_years']
    common_start = max(smb_years[0], d_years[0])
    common_end = min(smb_years[-1], d_years[-1])

    smb_mask = (smb_years >= common_start) & (smb_years <= common_end)
    d_mask = (d_years >= common_start) & (d_years <= common_end)
    years = smb_years[smb_mask]
    smb = data['smb'][:, smb_mask]  # (n_gl, n_years)
    d = data['d'][:, d_mask]        # (n_gl, n_years)

    n_years = len(years)
    areas = data['areas']

    # ── Per-glacier metrics ──
    early_mask = (years >= period_early[0]) & (years <= period_early[1])
    late_mask = (years >= period_late[0]) & (years <= period_late[1])

    cum_smb_deficit = np.full(n_gl, np.nan)
    d_change = np.full(n_gl, np.nan)
    d_accel = np.full(n_gl, np.nan)
    cum_mb = np.full(n_gl, np.nan)
    mean_d_early = np.full(n_gl, np.nan)

    for k in range(n_gl):
        smb_k = smb[k]
        d_k = d[k]

        if np.sum(np.isfinite(smb_k)) < 10 or np.sum(np.isfinite(d_k)) < 10:
            continue

        # Cumulative SMB anomaly relative to early period mean
        smb_early_mean = np.nanmean(smb_k[early_mask])
        smb_anom = smb_k - smb_early_mean
        cum_smb_deficit[k] = np.nansum(smb_anom)  # negative = deficit

        # D change: late mean minus early mean
        d_early = np.nanmean(d_k[early_mask])
        d_late = np.nanmean(d_k[late_mask])
        d_change[k] = d_late - d_early
        mean_d_early[k] = d_early

        # D acceleration: linear trend over full record
        valid = np.isfinite(d_k)
        if valid.sum() > 5:
            slope, _, _, _, _ = stats.linregress(years[valid], d_k[valid])
            d_accel[k] = slope  # Gt/yr²

        # Cumulative MB = cumulative(SMB - D)
        mb_k = smb_k - d_k
        cum_mb[k] = np.nansum(mb_k)

    # Filter to glaciers with valid data
    valid = (np.isfinite(cum_smb_deficit) & np.isfinite(d_change)
             & np.isfinite(d_accel) & np.isfinite(areas) & (areas > 0))
    n_valid = valid.sum()

    # Normalize by glacier area for fair comparison
    # SMB deficit per unit area (m w.e.) and D change per unit area
    GT_TO_M_WE_PER_KM2 = 1.0 / 1e-3  # 1 Gt = 1 km³ w.e. = 1e9 m³;
    # per km² = 1e9/1e6 = 1e3 m w.e. ... actually:
    # 1 Gt = 1e12 kg; area in km² = 1e6 m²; density water = 1000 kg/m³
    # thickness = mass / (density * area) = 1e12 / (1000 * 1e6) = 1 m w.e. per Gt per km²
    cum_smb_deficit_specific = cum_smb_deficit / areas  # m w.e.
    d_change_specific = d_change / areas                # Gt/yr/km²
    d_accel_specific = d_accel / areas

    # ── Statistics ──
    r_deficit_dchange, p_deficit_dchange = stats.pearsonr(
        cum_smb_deficit_specific[valid], d_change_specific[valid])
    r_deficit_daccel, p_deficit_daccel = stats.pearsonr(
        cum_smb_deficit_specific[valid], d_accel_specific[valid])
    r_cummb_dchange, p_cummb_dchange = stats.pearsonr(
        cum_mb[valid] / areas[valid], d_change_specific[valid])

    # ── Figure ──
    fig, axes = plt.subplots(2, 2, figsize=figsize)

    # Panel A: Cumulative SMB deficit vs D change (specific)
    ax = axes[0, 0]
    sc = ax.scatter(cum_smb_deficit_specific[valid], d_change_specific[valid],
                    c=areas[valid], cmap='viridis_r', s=20, alpha=0.7,
                    norm=plt.matplotlib.colors.LogNorm())
    ax.axhline(0, color='k', lw=0.5)
    ax.axvline(0, color='k', lw=0.5)
    ax.set_xlabel('Cumulative SMB deficit (m w.e.)')
    ax.set_ylabel('D change: late − early (Gt/yr/km²)')
    ax.set_title(f'SMB deficit vs discharge change\n'
                 f'r = {r_deficit_dchange:.3f}, p = {p_deficit_dchange:.3f}')
    ax.grid(True, alpha=0.3)
    cb = fig.colorbar(sc, ax=ax)
    cb.set_label('Glacier area (km²)')

    # Panel B: Cumulative SMB deficit vs D acceleration (specific)
    ax = axes[0, 1]
    ax.scatter(cum_smb_deficit_specific[valid], d_accel_specific[valid],
               c=areas[valid], cmap='viridis_r', s=20, alpha=0.7,
               norm=plt.matplotlib.colors.LogNorm())
    ax.axhline(0, color='k', lw=0.5)
    ax.axvline(0, color='k', lw=0.5)
    ax.set_xlabel('Cumulative SMB deficit (m w.e.)')
    ax.set_ylabel('D acceleration (Gt/yr²/km²)')
    ax.set_title(f'SMB deficit vs discharge acceleration\n'
                 f'r = {r_deficit_daccel:.3f}, p = {p_deficit_daccel:.3f}')
    ax.grid(True, alpha=0.3)

    # Panel C: Cumulative MB (SMB-D) vs D change
    ax = axes[1, 0]
    ax.scatter(cum_mb[valid] / areas[valid], d_change_specific[valid],
               c=areas[valid], cmap='viridis_r', s=20, alpha=0.7,
               norm=plt.matplotlib.colors.LogNorm())
    ax.axhline(0, color='k', lw=0.5)
    ax.axvline(0, color='k', lw=0.5)
    ax.set_xlabel('Cumulative MB (m w.e., negative = thinning)')
    ax.set_ylabel('D change: late − early (Gt/yr/km²)')
    ax.set_title(f'Total thinning vs discharge change\n'
                 f'r = {r_cummb_dchange:.3f}, p = {p_cummb_dchange:.3f}')
    ax.grid(True, alpha=0.3)

    # Panel D: Summary text
    ax = axes[1, 1]
    ax.axis('off')
    summary = (
        f"Per-glacier SMB–discharge feedback analysis\n"
        f"{'─' * 45}\n"
        f"Mouginot et al. (2019), marine-terminating only\n"
        f"N glaciers: {n_valid} (of {n_gl} TW)\n"
        f"Period: {common_start}–{common_end}\n"
        f"Early: {period_early[0]}–{period_early[1]}, "
        f"Late: {period_late[0]}–{period_late[1]}\n\n"
        f"Correlations (area-normalized):\n"
        f"  SMB deficit vs D change:  r={r_deficit_dchange:+.3f}  "
        f"p={p_deficit_dchange:.4f}\n"
        f"  SMB deficit vs D accel:   r={r_deficit_daccel:+.3f}  "
        f"p={p_deficit_daccel:.4f}\n"
        f"  Cum. MB vs D change:      r={r_cummb_dchange:+.3f}  "
        f"p={p_cummb_dchange:.4f}\n\n"
        f"Hypothesis: if SMB-driven thinning feeds back\n"
        f"on discharge via driving stress, glaciers with\n"
        f"larger SMB deficits should show more D increase.\n"
        f"This requires negative r (more deficit → more D)."
    )
    ax.text(0.05, 0.95, summary, transform=ax.transAxes,
            fontsize=10, verticalalignment='top', fontfamily='monospace')

    plt.tight_layout()

    if savefig:
        fig.savefig(savefig, dpi=150, bbox_inches='tight')
        print(f'Saved: {savefig}')

    results = {
        'n_valid': n_valid,
        'r_deficit_dchange': r_deficit_dchange,
        'p_deficit_dchange': p_deficit_dchange,
        'r_deficit_daccel': r_deficit_daccel,
        'p_deficit_daccel': p_deficit_daccel,
        'r_cummb_dchange': r_cummb_dchange,
        'p_cummb_dchange': p_cummb_dchange,
        'cum_smb_deficit_specific': cum_smb_deficit_specific,
        'd_change_specific': d_change_specific,
        'd_accel_specific': d_accel_specific,
        'cum_mb_specific': cum_mb / areas,
        'names': data['names'],
        'areas': areas,
        'valid': valid,
    }

    return fig, results


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        '../data/raw/ice_sheets/greenland/mouginot2019_data.xlsx'
    fig, results = analyze_smb_discharge_feedback(
        path,
        savefig='../figures/diagnostic_smb_discharge_feedback.png',
    )
    plt.show()
